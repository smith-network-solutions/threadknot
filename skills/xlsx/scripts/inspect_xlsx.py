#!/usr/bin/env -S uv run --quiet --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Print the structure and contents of an .xlsx workbook.

Run this before editing any workbook you did not create.

    inspect_xlsx.py budget.xlsx                      # every sheet, structure + preview
    inspect_xlsx.py budget.xlsx --sheet Q1           # one sheet
    inspect_xlsx.py budget.xlsx --range A1:F20       # a window
    inspect_xlsx.py budget.xlsx --formulas           # every formula cell
    inspect_xlsx.py budget.xlsx --values             # cached values, not formulas
    inspect_xlsx.py budget.xlsx --csv --sheet Q1     # CSV to stdout

Named `inspect_xlsx` rather than `inspect` on purpose: a module called
`inspect.py` shadows the standard library one that openpyxl's dependencies
import, and the failure looks like an unrelated circular import.
"""
import argparse
import csv
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


def fmt(value, width=18):
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\n", "\\n")
    return text if len(text) <= width else text[: width - 1] + "…"


def preview(ws, cell_range, max_rows, max_cols):
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_col, min_row = 1, 1
        max_col = min(ws.max_column, max_cols)
        max_row = min(ws.max_row, max_rows)
    max_col = min(max_col or 1, min_col + max_cols - 1)
    max_row = min(max_row or 1, min_row + max_rows - 1)

    header = "      " + " ".join(f"{get_column_letter(c):>18}" for c in range(min_col, max_col + 1))
    print(header)
    for r in range(min_row, max_row + 1):
        cells = []
        for c in range(min_col, max_col + 1):
            cells.append(f"{fmt(ws.cell(row=r, column=c).value):>18}")
        print(f"{r:>5} " + " ".join(cells))
    if ws.max_row > max_row:
        print(f"      ... {ws.max_row - max_row} more row(s)")
    if ws.max_column > max_col:
        print(f"      ... {ws.max_column - max_col} more column(s)")


def describe_sheet(ws, args):
    print(f"\n=== sheet {ws.title!r} ===")
    print(f"dimensions: {ws.dimensions}  ({ws.max_row} rows x {ws.max_column} cols)")
    if ws.freeze_panes:
        print(f"freeze panes: {ws.freeze_panes}")
    if ws.auto_filter and ws.auto_filter.ref:
        print(f"auto filter: {ws.auto_filter.ref}")
    if ws.sheet_state != "visible":
        print(f"state: {ws.sheet_state}")

    merged = [str(r) for r in ws.merged_cells.ranges]
    if merged:
        print(f"merged: {', '.join(merged[:12])}" + (" ..." if len(merged) > 12 else ""))

    widths = {k: v.width for k, v in ws.column_dimensions.items() if v.width}
    if widths:
        print("column widths: " + ", ".join(f"{k}={v:.0f}" for k, v in sorted(widths.items())[:12]))

    try:
        cf_ranges = list(ws.conditional_formatting)
    except Exception:
        cf_ranges = []
    if cf_ranges:
        print(f"conditional formatting on: {', '.join(str(c.sqref) for c in cf_ranges)}")

    if getattr(ws, "_charts", None):
        for chart in ws._charts:
            title = getattr(chart, "title", None)
            print(f"chart: {type(chart).__name__} {title if isinstance(title, str) else ''}")

    print("\npreview:")
    preview(ws, args.range, args.max_rows, args.max_cols)


def dump_formulas(wb):
    total = 0
    for ws in wb.worksheets:
        found = []
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    found.append((cell.coordinate, cell.value))
        if found:
            print(f"\n=== {ws.title} — {len(found)} formula cell(s) ===")
            for coord, formula in found[:200]:
                print(f"  {coord:>8}  {formula}")
            if len(found) > 200:
                print(f"  ... {len(found) - 200} more")
        total += len(found)
    if not total:
        print("no formulas in this workbook")


def dump_csv(ws):
    writer = csv.writer(sys.stdout)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if v is None else v for v in row])


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path")
    ap.add_argument("--sheet", help="limit to one sheet by name")
    ap.add_argument("--range", help="cell range to preview, e.g. A1:F20")
    ap.add_argument("--formulas", action="store_true", help="list every formula cell")
    ap.add_argument("--values", action="store_true",
                    help="show cached values instead of formulas (None until recalculated)")
    ap.add_argument("--csv", action="store_true", help="emit one sheet as CSV on stdout")
    ap.add_argument("--max-rows", type=int, default=25)
    ap.add_argument("--max-cols", type=int, default=8)
    args = ap.parse_args()

    try:
        wb = load_workbook(args.path, data_only=args.values or args.csv)
    except Exception as exc:  # noqa: BLE001
        print(f"could not open {args.path}: {exc}", file=sys.stderr)
        return 1

    if args.sheet and args.sheet not in wb.sheetnames:
        print(f"no sheet named {args.sheet!r}. Sheets: {', '.join(wb.sheetnames)}", file=sys.stderr)
        return 1
    sheets = [wb[args.sheet]] if args.sheet else wb.worksheets

    if args.csv:
        if len(sheets) != 1:
            print("--csv needs --sheet NAME", file=sys.stderr)
            return 2
        dump_csv(sheets[0])
        return 0

    if args.formulas:
        dump_formulas(wb)
        return 0

    print(f"=== {args.path} ===")
    print(f"sheets: {', '.join(wb.sheetnames)}")
    if wb.defined_names:
        names = [f"{n}={d.value}" for n, d in wb.defined_names.items()]
        print(f"named ranges: {', '.join(names[:12])}")
    if args.values:
        print("(showing CACHED VALUES — a formula never opened by a spreadsheet app reads as blank;"
              " run recalc.py to populate them)")

    for ws in sheets:
        describe_sheet(ws, args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
