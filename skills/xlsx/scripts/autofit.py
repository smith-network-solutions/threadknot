#!/usr/bin/env -S uv run --quiet --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Size every column to its content.

openpyxl never auto-sizes columns, so a workbook it wrote shows `####` wherever
a number is wider than the default width — the single most common reason a
generated spreadsheet looks broken to the person who opens it.

    autofit.py budget.xlsx                    # in place, every sheet
    autofit.py budget.xlsx out.xlsx --max 60
"""
import argparse
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def display_width(value, number_format: str) -> int:
    """Roughly how wide this cell renders.

    A formula is measured by its result, which we do not have — so formulas are
    given a modest fixed allowance rather than the width of their source text,
    which would make an `=SUMIFS(...)` column absurdly wide.
    """
    if value is None:
        return 0
    if isinstance(value, str) and value.startswith("="):
        return 12
    if isinstance(value, (int, float)):
        text = f"{value:,.2f}" if "0.00" in number_format or "#,##0" in number_format else str(value)
        return len(text)
    return max(len(line) for line in str(value).split("\n"))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("src")
    ap.add_argument("dest", nargs="?")
    ap.add_argument("--min", type=int, default=8, help="minimum width (default 8)")
    ap.add_argument("--max", type=int, default=60, help="maximum width (default 60)")
    ap.add_argument("--padding", type=int, default=2)
    ap.add_argument("--sheet", help="limit to one sheet")
    args = ap.parse_args()

    try:
        wb = load_workbook(args.src)
    except Exception as exc:  # noqa: BLE001
        print(f"could not open {args.src}: {exc}", file=sys.stderr)
        return 1

    sheets = [wb[args.sheet]] if args.sheet else wb.worksheets
    for ws in sheets:
        widest: dict[int, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                w = display_width(cell.value, cell.number_format or "")
                if w > widest.get(cell.column, 0):
                    widest[cell.column] = w
        for col, width in widest.items():
            final = max(args.min, min(args.max, width + args.padding))
            ws.column_dimensions[get_column_letter(col)].width = final
        print(f"{ws.title}: sized {len(widest)} column(s)")

    dest = args.dest or args.src
    wb.save(dest)
    print(f"wrote {dest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
